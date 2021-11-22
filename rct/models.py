#from flask_login import UserMixin

from flask import Flask
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import create_engine, Column, Integer, String, ForeignKey, Date, Boolean, PrimaryKeyConstraint, UniqueConstraint
from sqlalchemy.orm import relationship
from datetime import datetime
import openpyxl
import ipaddress
import string
import re
from transliterate import translit

db = SQLAlchemy()




class TimestampMixin:
    created_at = db.Column(db.TIMESTAMP, # тип данных,
        default = datetime.utcnow,
        nullable=False
        )
    updated_at = db.Column(db.TIMESTAMP,
        default = datetime.utcnow,      # вычисляется при создании
        onupdate=datetime.utcnow,        # вычисляется каждый раз при обновлении сущности
        nullable=False      # не может быть пустым
        )



class Address(db.Model, TimestampMixin):
    id = db.Column(db.Integer, primary_key=True)
    small_address = db.Column(db.String(50), comment='small_address')
    rgis_address  = db.Column(db.String(200), comment='rgis')
    district = db.Column(db.String(50), default='Красносельский', comment='rayon')
    translate = db.Column(db.String(50), comment='translate')
    places = db.relationship('Place', backref='placeaddr')
    vlans = db.relationship('Vlan', backref='addressvlans')
    switches = db.relationship('Switch', backref='addressswitch')
    vlanssw = db.relationship('Vlansw', backref='addressvlanssw')
    nodes = db.relationship('Node', backref='addressnodes', uselist = False)
    ports = db.relationship('Port', backref='addressports')


class Vlansw(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_vl = db.Column(db.Integer, nullable=False, comment='id_vlan', unique=True)
    name = db.Column(db.String(14), nullable=False, comment='name_vlan', unique=True)
    ipnet = db.Column(db.String(20), comment='ip_network_mng', unique=True)
    netmask = db.Column(db.Integer, comment='network_mask')
    gw = db.Column(db.Integer, comment='network_mask') # net4.network_address+1
    address_id = db.Column(db.Integer, db.ForeignKey('address.id'), comment='Адрес узла')
    ipaddrsw = relationship("Ipaddrsw", backref="vlnetswipaddr",cascade="all,delete", passive_deletes=True)

class Location(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    small_location = db.Column(db.String(10),unique=True)
    places = db.relationship('Place', backref='placelocat')


class Type(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    type = db.Column(db.String(7),unique=True)
    vlans = db.relationship('Vlan', backref='typevlans')
    devises = relationship("Device", backref="typedevice")

class Typesw(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    typesw = db.Column(db.String(7), nullable = False, unique=True)
    #vlans = db.relationship('Vlan', backref='typevlans')
    switches = relationship("Switch", backref="typeswitch")


class Place(db.Model, TimestampMixin):
    #__table_args__ = (PrimaryKeyConstraint('location_id', 'address_id'),)
    __table_args__ = (UniqueConstraint('location_id', 'address_id', name ='place_address'),)
    id = db.Column(db.Integer, primary_key=True)
    guid = db.Column(db.String(50), default='no', comment='name')
    id_gmc = db.Column(db.String(10), default='no', comment='ID_GMC')
    description_gmc = db.Column(db.String(10), default='no', comment='description-GMC')
    lng = db.Column(db.Integer, default='0', comment='long')
    lat = db.Column(db.Integer, default='0', comment='lat')
    rtsp = db.Column(db.String(40), default='no', comment='rtsp')
    cam_type = db.Column(db.String(40), default='no', comment='type')
    devices = db.relationship('Device', backref = 'place', uselist = False)
    switches = db.relationship('Switch', backref = 'place', uselist = False)
    address_id = db.Column(db.Integer, db.ForeignKey('address.id'), nullable=False)
    location_id = db.Column(db.Integer, db.ForeignKey('location.id'), nullable=False)
    #port_id = db.Column(db.Integer, db.ForeignKey('port.id'))
    port = db.relationship('Port', backref='placeport', uselist = False)
    azimuth = db.Column(db.String(15), default='no', comment='azimut')
    angle = db.Column(db.String(15), default='no', comment='angel')
    n_gk = db.Column(db.String(15), default='GK-', comment='N_GK')
    osd_menu = db.Column(db.String(15), default='OSD', comment='osd')
    screen_pth_v1 = db.Column(db.String(100), default='no', comment='screen_first')
    screen_pth_v2 = db.Column(db.String(100), default='no', comment='screen_last')
    vlan_id = db.Column(db.Integer, db.ForeignKey('vlan.id'), nullable=False)
    active = db.Column(db.Boolean, default = True, nullable = False, comment = 'ONLINE или OFFLINE')

class Server(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(40), default='vds-', comment='name_server')
    ipaddr_id = db.Column(db.Integer, db.ForeignKey('ipaddr.id'))
    devices = db.relationship('Device', backref='deviceserver')

# Под вопросом
class Ipaddr(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ipaddr = db.Column(db.String(26), default='0.0.0.0', comment='ipaddr')
    devices = db.relationship('Device', backref='deviceip', uselist = False)
    servers = db.relationship('Server', backref='serverip', uselist = False)

# под вопросом


class Vlan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    id_vl = db.Column(db.Integer, nullable=False, comment='id_vlan', unique=True)
    name = db.Column(db.String(14), nullable=False, comment='name_vlan', unique=True)
    ipnet = db.Column(db.String(20), comment='ip_network', unique=True)
    netmask = db.Column(db.Integer, comment='network_mask')
    placeses = db.relationship('Place', backref = 'vlanplace')
    address_id = db.Column(db.Integer, db.ForeignKey('address.id'))
    type_id = db.Column(db.Integer, db.ForeignKey('type.id'))
    model_id = db.Column(db.Integer, db.ForeignKey('model.id'))

class Ipaddrsw(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    ipaddr = db.Column(db.String(26), default='2.2.2.2', comment='ipaddrsw')
    netmask = db.Column(db.Integer, comment='network_mask', default = Vlan.netmask)
    gw = db.Column(db.String(26), default='2.2.2.2', comment='gw')
    status = db.Column(db.String(26), default='Free', comment='Free or Busy')
    switches = db.relationship('Switch', backref='switchip')
    vlansw_id = db.Column(db.Integer, db.ForeignKey('vlansw.id', ondelete='CASCADE'))

    @classmethod
    def createips(cls,objvlnt):
        net4 = ipaddress.ip_network(str(objvlnt.ipnet) + '/' + str(objvlnt.netmask))
        list_ipaddr = [str(ip) for ip in net4.hosts() if ip != net4.network_address+1]

        for ip in list_ipaddr:
            dict = {key: item for key, item in zip(['ipaddr', 'netmask', 'gw', 'status', 'vlansw_id'],
                                                   [ip, objvlnt.netmask, objvlnt.gw, 'free', objvlnt.id])}
            ipitem = Ipaddrsw(**dict)
            db.session.add(ipitem)
            db.session.flush()
            db.session.commit()



class Device(db.Model, TimestampMixin):
    id = db.Column(db.Integer, primary_key=True)
    id_aiu = db.Column(db.String(10), comment='ID_AIU', nullable=False, unique=True)
    mac = db.Column(db.String(16), comment='mac', unique=True)
    #ip = db.Column(db.String(10), comment='ip', default = '192.168.1.64')
    #mask = db.Column(db.String(2), comment='mask', default = '24')
    docs = db.Column(db.String(30), comment='Накладная')
    type_id = db.Column(db.Integer, db.ForeignKey('type.id'))
    model_id = db.Column(db.Integer, db.ForeignKey('model.id'))
    place_id = db.Column(db.Integer, db.ForeignKey('place.id'))
    ipaddr_id = db.Column(db.Integer, db.ForeignKey('ipaddr.id'),  unique=True)
    server_id = db.Column(db.Integer, db.ForeignKey('server.id'))

    ######### mac


class Model(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    model = db.Column(db.String(40), unique=True)
    vendor = db.Column(db.String(40), comment='vendor')  # необходимо вычислять его автоматически
    devices = relationship("Device", backref="modeldevice")


    def __init__(self, model, vendor):
        self.model = model
        if 'DS-2' in str(model):
            vendor = "hikvision"
        elif 'VTO' in str(model):
            vendor = "dahua"
        self.vendor = vendor

    def __repr__(self):
        return '<Model %r>' % self.model







# Предназначен для хранение основных параметров коммутатора, чтобы в автоматическом режиме создавать порты
class Modelswitch(db.Model):
    #__tablename__ = 'ModelSwitch'
    id = db.Column(db.Integer, primary_key=True)
    modelsw = db.Column(db.String(10),unique=True, nullable=False, comment = 'Название модели')
    qt_port = db.Column(db.Integer, default = 'None',comment = 'Кол-во портов всего')
    list_combo_port =db.Column(db.String(50),default = 'None', comment = 'Список combo портов')
    list_poe = db.Column(db.String(200),default = 'None', comment = 'Список poe портов')
    list_eth = db.Column(db.String(40),default = 'None', comment='Список Ethernet портов')
    list_cx = db.Column(db.String(40),default = 'None', comment = 'Список CX портов')
    list_sfp = db.Column(db.String(200),default = 'None', comment = 'Список SFP портов')
    list_sfp_plus = db.Column(db.String(200),default = 'None', comment='Список SFP+ портов')
    switches = relationship("Switch", backref="modelswitch")


# Узел
class Node(db.Model, TimestampMixin):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(25), comment='Имя узла', nullable = False, unique=True)
    description = db.Column(db.String(400), comment='Примечание')
    address_id = db.Column(db.Integer, db.ForeignKey('address.id'), unique=True)
    switches = db.relationship('Switch', backref='swnodes')




class Switch(db.Model, TimestampMixin):
    id = db.Column(db.Integer, primary_key=True)
    id_aiu = db.Column(db.String(10), comment='ID_AIU', nullable=False, unique=True)
    name = db.Column(db.String(10), unique=True)
    mac = db.Column(db.String(16), comment='mac', unique=True)
    docs = db.Column(db.String(30), comment='Накладная')
    description = db.Column(db.String(400), comment='Примечание')
    type_id = db.Column(db.Integer, db.ForeignKey('typesw.id'))
    model_id = db.Column(db.Integer, db.ForeignKey('modelswitch.id'))
    place_id = db.Column(db.Integer, db.ForeignKey('place.id'))  # Удалить
    address_id = db.Column(db.Integer, db.ForeignKey('address.id'))
    ipaddrsw_id = db.Column(db.Integer, db.ForeignKey('ipaddrsw.id'))
    node_id = db.Column(db.Integer, db.ForeignKey('node.id'))
    ports = db.relationship('Port', backref='swports', cascade="all,delete", passive_deletes=True)




class Port(db.Model, TimestampMixin):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.Integer)
    switch_id = db.Column(db.Integer, db.ForeignKey('switch.id', ondelete='CASCADE')) # коммутатор которому принадлежит self порт
    status = db.Column(db.String(10), default = 'off', comment = 'empty')
    linksw_id = db.Column(db.Integer, db.ForeignKey('port.id')) # подключенный к порту коммутатор
    linkdev_id =db.Column(db.Integer, db.ForeignKey('place.id')) # подключенный к порту device
    address_id = db.Column(db.Integer, db.ForeignKey('address.id'))
    linkssw = relationship("Port")
    description = db.Column(db.String(400), comment='Примечание')

    @classmethod
    def createports(cls, objsw):
        #
        # qtport = objsw.modelswitch.qt_port
        # print('Число портов',qtport)
        print('все атрибуты',objsw.__dict__)

        for portname in range(objsw.modelswitch.qt_port):
            #  dict = {key: item for key, item in zip(['name', 'switch_id', 'status', 'linksw_id', 'linkdev_id'],
            #                                                    ['Port_'+ str(portname + 1), objsw.id, 'free', 'free', 'free'])}
            dict = {key: item for key, item in zip(['name', 'switch_id', 'status'],
                                                   ['Port_'+ str(portname + 1), objsw.id])}

            port = Port(**dict)
            db.session.add(port)
            db.session.flush()
            db.session.commit()




def vlanFromFile(filename, app):
    source_excel = app.config['UPLOAD_FOLDER']+'/' + filename
    wb = openpyxl.load_workbook(source_excel)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    status_list = []

    head_dict = {
        "VLAN ID": "VLAN ID",
        "Имя VLAN": "Имя VLAN",
        "Адрес" : "Адрес",
        "IP СЕТИ" : "IP СЕТИ" ,
        "Mask": "Mask",
        "Тип": "Тип",
        "Статус" : "Статус"}

    status_list.append(head_dict)

    for i in range(2, rows + 1):
        status_dict={}
        id_vl = sheet.cell(row=i, column=1).value
        status_dict["vlan_id"] = id_vl
        name = sheet.cell(row=i, column=2).value
        status_dict["name"] = name
        type = sheet.cell(row=i, column=3).value
        small_address = (sheet.cell(row=i, column=4).value).strip()
        ipnet = (sheet.cell(row=i, column=5).value).strip()
        netmask = (sheet.cell(row=i, column=6).value)





        small_address = small_address.strip()

        addr = Address.query.filter(Address.small_address.contains(small_address)).first()
        if addr and check_if_ip_is_network(ipnet, netmask):
            status_dict['address'] = addr.small_address
            status_dict['ipnet'] = ipnet
            status_dict['netmask'] = netmask
            type = Type.query.filter(Type.type.contains(type)).first()
            if type:

                status_dict['type'] = type.type

                try:
                    vlan = Vlan(id_vl=id_vl, name=name, address_id=addr.id, type_id=type.id, ipnet = ipnet, netmask = netmask)
                    db.session.add(vlan)
                    db.session.flush()
                    db.session.commit()

                    status_dict['status'] = "Добавлено в базу"
                except:
                    db.session.rollback()
                    status_dict['status'] = "Ошибка добавления"
        else:
            status_dict['address'] = small_address
            status_dict['ipnet'] = check_if_ip_is_network(ipnet, netmask)
            status_dict['mask'] = check_if_ip_is_network(ipnet, netmask)
            status_dict['type'] = type
            status_dict['status'] = "Необходимо добавить адрес в базу"
        status_list.append(status_dict)

    return status_list

def modeliFromFile(filename, app):
    source_excel = app.config['UPLOAD_FOLDER'] + '/' + filename
    wb = openpyxl.load_workbook(source_excel)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    status_list = []

    head_dict = {
        "Модель": "Модель",
        "Статус": "Статус"
    }

    status_list.append(head_dict)

    for i in range(2, rows + 1):
        status_dict = {}

        model = sheet.cell(row=i, column=1).value
        model = runupmodel(model)
        status_dict["modeli"] = model

        modelstatus = Model.query.filter(Model.model.contains(model)).first()
        if modelstatus:
            status_dict['status'] = 'Модель уже существует'

        else:
           try:
               model = Model(model, vendor="unknow")
               db.session.add(model)

               db.session.flush()
               db.session.commit()
               status_dict['status'] = "Добавлено в базу"
           except:
               db.session.rollback()
               status_dict['status'] = "Ошибка добавления"

        status_list.append(status_dict)

    return status_list


# Загрузка моделей коммутаторов из файла
def modeliswFromFile(filename, app):
    source_excel = app.config['UPLOAD_FOLDER'] + '/' + filename
    wb = openpyxl.load_workbook(source_excel)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    status_list = []

    head_dict = {
        "Модель": "Модель",
        "Статус": "Статус"
    }

    status_list.append(head_dict)

    for i in range(2, rows + 1):
        status_dict = {}

        modelsw = sheet.cell(row=i, column=1).value
        modelsw = runupmodel(modelsw)  # Пропускаем через функцию
        status_dict["modelisw"] = modelsw

        modelstatus = Modelswitch.query.filter(Modelswitch.modelsw.contains(modelsw)).first()
        if modelstatus:
            status_dict['status'] = 'Модель уже существует'

        else:
           try:

               qt_port = sheet.cell(row=i, column=2).value
               list_combo_port = sheet.cell(row=i, column=3).value
               list_poe = sheet.cell(row=i, column=4).value
               list_eth = sheet.cell(row=i, column=5).value
               list_cx = sheet.cell(row=i, column=6).value
               list_sfp = sheet.cell(row=i, column=7).value
               list_sfp_plus = sheet.cell(row=i, column=8).value

               modelsw = Modelswitch(modelsw = modelsw, qt_port = qt_port, list_combo_port = list_combo_port, list_poe = list_poe, list_eth = list_eth, list_cx = list_cx, list_sfp = list_sfp, list_sfp_plus=list_sfp_plus)
               db.session.add(modelsw)

               db.session.flush()
               db.session.commit()
               status_dict['status'] = "Добавлено в базу"
           except:
               db.session.rollback()
               status_dict['status'] = "Ошибка добавления"

        status_list.append(status_dict)

    return status_list5


def deviceFromFile(filename, app):
    source_excel = app.config['UPLOAD_FOLDER'] + '/' + filename
    wb = openpyxl.load_workbook(source_excel)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    status_list = []

    head_dict = {
        "Id_АИЮ": "Id_АИЮ",
        "Mac": "Mac",
        "Накладная":"Накладная",
        "Тип": "Тип",
        "Модель": "Модель",
        "status": "Статус"
    }

    status_list.append(head_dict)

    for i in range(2, rows + 1):
        print(i)
        status_dict = {}


        # ID DEVICE  NO: преобразование проверка и тд
        id_aiu = sheet.cell(row=i, column=1).value
        if id_aiu:
            id_aiustatus = Device.query.filter(Device.id_aiu.contains(id_aiu)).first()
            # Проверка: если ID уществует, то нет смысла идти дальше
            if id_aiustatus:

                status_dict["Id_АИЮ"] = id_aiu
                status_dict["Тип"] = '---'
                status_dict["mac"] = '---'
                status_dict["Модель"] = '---'
                status_dict["Накладная"] = '---'
                status_dict["status"] = 'Такой ID уже существует'


            else:
                status_dict["Id_АИЮ"] = id_aiu

                # Работа с mac
                mac =  sheet.cell(row=i, column=2).value
                mac = check_if_mac_aiu(mac)
                status_dict["Mac"] = mac

                # Работа с Накладной
                docs = sheet.cell(row=i, column=3).value
                status_dict["Накладная"] = docs

                # Работа с типом устройства и моделью устройства => ОБЯЗАТЕЛЬНО
                type = (sheet.cell(row=i, column=4).value).strip()
                type = Type.query.filter(Type.type.contains(type)).first()
                model = (sheet.cell(row=i, column=5).value).strip()

                model = Model.query.filter(Model.model.contains(model)).first()
                if type:
                    status_dict["Тип"] = type.type

                    if model:
                        status_dict["Модель"] = model.model

                        try:
                            status_dict["Тип"] = type.type
                            status_dict["Накладная"] = docs
                            dev = Device(id_aiu=id_aiu, mac=mac, docs=docs, type_id=type.id, model_id=model.id)
                            db.session.add(dev)
                            db.session.flush()
                            db.session.commit()


                            status_dict["status"] = "Успешно добавлено В БД"
                        except:
                            db.session.rollback()
                            status_dict["status"] = "Ошибка добавления"

                    else:

                        status_dict["Тип"] = '---'
                        status_dict["Модель"] = (sheet.cell(row=i, column=5).value).strip()
                        status_dict["Накладная"] = '---'
                        status_dict["status"] = "Добавьте модель в базу"


                else:
                    status_dict["Тип"] = 'ошибка'
                    status_dict["status"] = 'неверно указан тип'
                    status_dict["mac"] = '---'
                    status_dict["Накладная"] = '---'
                    status_dict["Модель"] = '---'

        # Добавление в список словаря по каждому экземпляру (по строке из файла)
        status_list.append(status_dict)

    return status_list



def locatFromFile(filename, app):
    source_excel = app.config['UPLOAD_FOLDER'] + '/' + filename
    wb = openpyxl.load_workbook(source_excel)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    status_list = []

    head_dict = {
        "Локация": "Локация",
        "Статус": "Статус"
    }

    status_list.append(head_dict)

    for i in range(2, rows + 1):
        status_dict = {}
        small_location = sheet.cell(row=i, column=1).value
        status_dict["small_location"] = small_location

        locat = Location.query.filter(Location.small_location.contains(small_location)).first()


        if locat:
            status_dict['status'] = 'Локация уже существует'

        else:
           try:
                location = Location(small_location=small_location)
                db.session.add(location)
                db.session.flush()
                db.session.commit()
                status_dict['status'] = "Добавлено в базу"
           except:
               db.session.rollback()
               status_dict['status'] = "Ошибка добавления"

        status_list.append(status_dict)

    return status_list


def addrFromFile(filename, app, translit):
    source_excel = app.config['UPLOAD_FOLDER']+'/' + filename
    wb = openpyxl.load_workbook(source_excel)
    sheet = wb.active
    rows = sheet.max_row
    cols = sheet.max_column
    status_list = []

    head_dict = {
        "АДРЕС" : "АДРЕС",
        "Адрес РГИС" : "Адрес РГИС",
        "Район" : "Район",
        "Статус" : "Статус"}

    status_list.append(head_dict)

    for i in range(2, rows + 1):
        status_dict = {}
        small_address = sheet.cell(row=i, column=1).value
        status_dict["small_address"] = small_address
        rgis_address = sheet.cell(row=i, column=2).value
        status_dict["rgis_address"] = rgis_address
        district  = sheet.cell(row=i, column=3).value
        status_dict["district"] = district

        small_address = runupaddr(small_address)

        rgis_address = rgis_address.strip()
        translate = translit(small_address, language_code='ru', reversed=True)
        translate = translate.replace("'","")
        addr = Address.query.filter(Address.small_address.contains(small_address)).first()
        addrrgis = Address.query.filter(Address.rgis_address.contains(rgis_address)).first()

        if addr and addrrgis:
            status_dict['status'] = 'ПАРА small adsress и rgis уже существуют'

        else:
           try:
                address = Address(small_address=small_address, rgis_address = rgis_address, district=district, translate = translate)
                db.session.add(address)
                db.session.flush()
                db.session.commit()
                status_dict['status'] = "Добавлено в базу"
           except:
               db.session.rollback()
               status_dict['status'] = "Ошибка добавления"

        status_list.append(status_dict)

    return status_list
#####
                    # СДЕЛАНО через @classmethod
                    # # Автосоздание ip адресов vlnt - vlans_network obj
                    # def auto_create_ipaddr(vlnt):
                    #     print('vlan_id ',vlnt.id, 'ipnet ', vlnt.ipnet,'mask ', vlnt.netmask, 'gw ', vlnt.gw)
                    #     net4 = ipaddress.ip_network(str(vlnt.ipnet)+'/'+str(vlnt.netmask))
                    #     list_ipaddr = [str(ip) for ip in net4.hosts() if ip != net4.network_address+1]
                    #
                    #
                    #     for ip in list_ipaddr:
                    #         dict = {key : item for key, item in zip(['ipaddr', 'netmask', 'gw', 'status', 'vlansw_id'], [ip, vlnt.netmask, vlnt.gw, 'free', vlnt.id])}
                    #
                    #         #try:
                    #         ipitem = Ipaddrsw(**dict)
                    #         db.session.add(ipitem)
                    #         db.session.flush()
                    #         db.session.commit()
                    #
                    #         #except:
                    #             # db.session.rollback()
                    #             # print('Ошибка создания порта')
                    #
                    #     #list2 = [ip for {key : item for key, item in zip(['ipaddr', 'netmask', 'gw', 'status', 'vlansw_id'], [ip, vlnt.netmask, vlnt.gw, 'free', vlnt.id])} in net4.hosts() if ip != net4.network_address+1]
                    #
                    #     #print(list2)














def runupaddr(small_address):
    """Удаление пробелов и добавление заглавной буквы"""
    small_address = small_address.strip()
    small_address = small_address.title()
    return small_address

def runupmodel(model):
    """Удаление пробелов и все буквы заглавные"""
    model = model.strip()
    model = model.upper()
    return model

def check_if_id_aiu(id_aiu):
    """проверка ID AIU regex   \d{4}$ and  ВП-(\d{2} |\d{2} | \d{3})$"""
    print ('ID AIU',id_aiu)
    return id_aiu

def check_if_mac_aiu(mac):
    """проверка mac и возврат к 00:00:00"""
    # Удяляем пробелы и делаем все буквы заглавными
    if mac:
        mac = mac.strip()
        # Приводим к виду :
        macd = ''
        mac_re = re.compile(r'\w{2}')
        mac_i = mac_re.findall(mac)
        for i in mac_i:
            macd = macd + str(i) + ':'
        mac = (macd[0:17]).lower()

        # переводим в верхней регистр
        mac = mac.upper()
        mac_re = re.compile(r'^([0-9A-F]{1,2})(\:[0-9A-F]{1,2}){5}$')
        if mac_re.search(mac) is not None:
            return mac
        else:
            return None
    return None


def check_if_ip_aiu(ip):
    """проверка ip  принадлежности его к сети, нужно передавать еще и место"""
    pass
    return ip

def check_addsw_ipaddrsw():
    """Проверка принадлженость адреса установки коммутатора"""


# проверка правильности введения сети и маски
def check_if_ip_is_network(ipnet, netmask):
    ip_address = str(ipnet) + '/' + str(netmask)
    try:
        ipaddress.ip_network(ip_address)
        return True
    except:
        return False

# def excel_to_dict():
#     """Делает из экселевского файла список словарей по строкам, где заголовки это ключи """
#     # считываем файл
#     source_excel = 'formexcel.xlsx'
#     #source_excel = app.config['UPLOAD_FOLDER'] + '/' + filename
#     wb = openpyxl.load_workbook(source_excel)
#     sheet = wb.active
#     rows = sheet.max_row
#     cols = sheet.max_column
#
#     # Создаем список, в который будем добавлять словари
#     status_list = []
#
#     # Считываем заголовки


